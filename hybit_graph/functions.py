from openpyxl import load_workbook
import pandas as pd

from hybit_graph.classes import Node, Edge

def load_colorcodes(path):
    code_dict = {}
    with open(path) as file:
        for line in file:
            content = line.rstrip('\n').split(',')
            print(content)
            key, value = int(content[0]), content[1]
            code_dict[key] = value
    return code_dict

def load_data_pd(filename):
    abrr_dict = {'wesernetz':'wn', 'AB 1.1 (Mess-Infrastuktur)':'ab1.1'}
    df = pd.read_excel(filename, header=3, usecols='B:D,L',)
    df.dropna(subset=['Senke'], inplace=True)
    df['Senke'] = df.Senke.apply(lambda x: str(x).rstrip(';').split(';'))
    df['woher-ext'] = df['woher-ext'].apply(lambda x: str(x).lower().replace(',', ';').split(';'))
    df = df.explode('Senke')
    df = df.explode('woher-ext')
    df['woher-ext'] = df['woher-ext'].apply(str.strip)
    df = df.loc[df['Senke'] != 'xxxxxxx']
    df = df.replace(abrr_dict)

    nodes = {}
    edges_weight = {}
    edges = []

    for index, row in df.iterrows():
        source_id = None
        sink_id = None
        #source
        if str(row['Quelle']) != 'nan':
            source_id = str(row['Quelle'])
        elif str(row['woher-int']) != 'nan':
            source_id = str(row['woher-int'])
        elif str(row['woher-ext']) != 'nan':
            source_id = str(row['woher-ext'])
        else:
            print("Missing source id in %s, line ignored" % index)
            continue
        source_id = str(source_id)
        if source_id not in nodes:
            nodes[source_id] = Node(source_id)
        # Senke
        if str(row['Senke']) != 'nan':
            sink_id = str(row['Senke'])
            if sink_id not in nodes:
                nodes[sink_id] = Node(sink_id)
        if source_id and sink_id:
            edge = Edge('', source_id, '', sink_id)
            edge.node_in = nodes[sink_id]
            edge.node_out = nodes[source_id]

            nodes[source_id].edges_out.append(edge)
            nodes[sink_id].edges_in.append(edge)
            edges.append(edge)
            if source_id not in edges_weight:
                    edges_weight[source_id] = {}

            if sink_id not in edges_weight[source_id]:
                edges_weight[source_id][sink_id] = 0

            edges_weight[source_id][sink_id] += 1
    return nodes, edges, edges_weight
            

def load_data(filename):
    wb = load_workbook(filename=filename)
    sheet = wb.active

    nodes = {}
    edges_weight = {}
    edges = []

    for row in sheet.iter_rows(min_row=5, max_col=6):
        source_id = None
        sink_id = None

        # source
        if row[2].value is not None and len(str(row[2].value)) > 0:
            # valid node
            source_id = str(row[2].value).strip()
            if source_id not in nodes:
                nodes[source_id] = Node(source_id)
        else:
            print("Missing source id in %s, line ignored" % row[2].coordinate)
            continue

        # sink
        if source_id and row[5].value is not None and len(str(row[5].value)) > 0:

            targets = str(row[5].value).split(";")

            for target in targets:
                # valid node
                sink_id = target.strip()
                if sink_id not in nodes:
                    nodes[sink_id] = Node(sink_id)

                if source_id and sink_id:
                    # add edge
                    edge = Edge(row[0].value, row[1].value, row[3].value, row[4].value)
                    edge.node_in = nodes[source_id]
                    edge.node_out = nodes[sink_id]

                    nodes[source_id].edges_out.append(edge)
                    nodes[sink_id].edges_in.append(edge)
                    edges.append(edge)
                    if source_id not in edges_weight:
                        edges_weight[source_id] = {}

                    if sink_id not in edges_weight[source_id]:
                        edges_weight[source_id][sink_id] = 0

                    edges_weight[source_id][sink_id] += 1
        else:
            print("Missing sink id in %s, line ignored" % row[5].coordinate)

    return nodes, edges, edges_weight
